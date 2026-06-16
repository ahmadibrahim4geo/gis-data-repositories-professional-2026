"""
Export Master_Index and Free_GIS_Data sheets from Excel to CSV and JSON.

Usage:
    python scripts/export_excel_to_csv_json.py
"""

import sys
import csv
import json
from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
EXCEL_PATH = DATA_DIR / "GIS_Data_Repositories_Professional_2026.xlsx"

SHEET_MAP = {
    "master_index": "Master_Index",
    "free_gis_data": "Free_GIS_Data",
}

OUTPUT_PATHS = {
    "master_index": {
        "csv": DATA_DIR / "master_index.csv",
        "json": DATA_DIR / "master_index.json",
    },
    "free_gis_data": {
        "csv": DATA_DIR / "free_gis_data.csv",
        "json": DATA_DIR / "free_gis_data.json",
    },
}


def clean_value(val):
    if val is None:
        return ""
    if isinstance(val, float) and (val != val):
        return ""
    return val


def export_sheet(wb, sheet_key):
    sheet_name = SHEET_MAP.get(sheet_key)
    if sheet_name not in wb.sheetnames:
        available = [s for s in wb.sheetnames if sheet_key.replace("_", " ").lower() in s.lower()]
        if available:
            sheet_name = available[0]
            print(f"  Sheet '{SHEET_MAP[sheet_key]}' not found, using '{sheet_name}' instead")
        else:
            print(f"  Skipping '{sheet_key}': no matching sheet found")
            return

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  Sheet '{sheet_name}' is empty")
        return

    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        record = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            record[key] = clean_value(val)
        data.append(record)

    csv_path = OUTPUT_PATHS[sheet_key]["csv"]
    json_path = OUTPUT_PATHS[sheet_key]["json"]

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(data)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"  Exported {len(data):,} records to:")
    print(f"    CSV: {csv_path.name}")
    print(f"    JSON: {json_path.name}")
    return len(data)


def main():
    print(f"GIS Data Repositories — Export Script")
    print(f"Running at: {datetime.now().isoformat()}")
    print(f"Excel path: {EXCEL_PATH}")
    print()

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at: {EXCEL_PATH}")
        print("Run this script from the project root directory.")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Workbook sheets: {len(wb.sheetnames)}")
    print()

    counts = {}
    for key in SHEET_MAP:
        print(f"--- {key} ---")
        count = export_sheet(wb, key)
        if count is not None:
            counts[key] = count
        print()

    wb.close()

    print("=" * 50)
    print("Export Summary:")
    for key, count in counts.items():
        csv_p = OUTPUT_PATHS[key]["csv"]
        json_p = OUTPUT_PATHS[key]["json"]
        print(f"  {key}: {count:,} records -> {csv_p.name}, {json_p.name}")
    print("Done.")


if __name__ == "__main__":
    main()
