"""
Rebuild dashboard statistics from actual workbook content.

Usage:
    python scripts/rebuild_dashboard_stats.py
"""

from datetime import datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
EXCEL_PATH = DATA_DIR / "GIS_Data_Repositories_Professional_2026.xlsx"


def main():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        return

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if "Dashboard" not in wb.sheetnames:
        print("Dashboard sheet not found")
        wb.close()
        return

    ws = wb["Dashboard"]
    print(f"Dashboard sheet: {ws.max_row} rows x {ws.max_column} cols")

    if "Master_Index" not in wb.sheetnames:
        print("Master_Index sheet not found — cannot calculate stats")
        wb.close()
        return

    mi = wb["Master_Index"]
    total_records = mi.max_row - 1

    status_col = None
    header_row = list(mi.iter_rows(min_row=1, max_row=1, values_only=True))
    if header_row:
        for i, h in enumerate(header_row[0]):
            if h and "status" in str(h).lower():
                status_col = i + 1
                break

    working = 0
    redirected = 0
    broken = 0
    needs_review = 0

    if status_col:
        for row in mi.iter_rows(min_row=2, max_row=mi.max_row, min_col=status_col, max_col=status_col, values_only=True):
            val = str(row[0]).strip() if row[0] else ""
            if val.lower() == "working":
                working += 1
            elif val.lower() == "redirected":
                redirected += 1
            elif "broken" in val.lower() or "dead" in val.lower() or "error" in val.lower():
                broken += 1
            elif val and val.lower() not in ("", "none"):
                needs_review += 1
            else:
                needs_review += 1
    else:
        print("  Status column not found in Master_Index")
        working = "Needs Review"
        redirected = "Needs Review"
        broken = "Needs Review"
        needs_review = "Needs Review"

    total_sheets = len(wb.sheetnames)
    wb.close()

    print()
    print("Dashboard Statistics (calculated):")
    print(f"  Total Sheets: {total_sheets}")
    print(f"  Total Records: {total_records}")
    print(f"  Working: {working}")
    print(f"  Redirected: {redirected}")
    print(f"  Broken: {broken}")
    print(f"  Needs Review: {needs_review}")
    print(f"  Generated: {datetime.now().isoformat()}")


if __name__ == "__main__":
    main()
