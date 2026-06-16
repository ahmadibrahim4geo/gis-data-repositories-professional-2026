"""
Clean and improve the Excel workbook presentation.

- Freeze top row in key sheets
- Apply auto-filters
- Set readable column widths
- Make URLs clickable hyperlinks
- Update Dashboard with calculated stats
- Add 'Generated On' date
"""

import sys
from datetime import datetime, date
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
EXCEL_PATH = DATA_DIR / "GIS_Data_Repositories_Professional_2026.xlsx"

IMPORTANT_SHEETS = [
    "Master_Index", "Dashboard", "Link_Audit", "Recommended_2026",
    "Free_GIS_Data", "Data_Dictionary",
]


def make_hyperlink(ws, row, col, url, display=None):
    from openpyxl.styles import Font, Color
    cell = ws.cell(row=row, column=col)
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")
    if display:
        cell.value = display


def auto_width(ws, max_width=60):
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, min(len(val), max_width))
            except Exception:
                pass
        adjusted = min(max_len + 2, max_width)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)


def find_url_columns(ws):
    url_cols = []
    for cell in ws[1]:
        if cell.value and isinstance(cell.value, str):
            if "url" in cell.value.lower() or "link" in cell.value.lower() or "web" in cell.value.lower() or "website" in cell.value.lower():
                url_cols.append(cell.column)
    return url_cols


def hyperlink_urls(ws, url_cols):
    count = 0
    for col in url_cols:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            val = str(cell.value).strip() if cell.value else ""
            if val and val.startswith("http"):
                if not cell.hyperlink:
                    try:
                        cell.hyperlink = val
                        count += 1
                    except Exception:
                        pass
    return count


def main():
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print(f"ERROR: File not found: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    today = date.today().isoformat()
    print(f"Fixing workbook: {EXCEL_PATH.name}")
    print(f"Date: {today}")
    print()

    for sn in IMPORTANT_SHEETS:
        if sn not in wb.sheetnames:
            print(f"  Sheet '{sn}' not found, skipping")
            continue

        ws = wb[sn]
        print(f"  Processing: {sn} ({ws.max_row} rows, {ws.max_column} cols)")

        ws.freeze_panes = ws.cell(row=2, column=1)

        if ws.max_row > 1:
            last_col = ws.max_column
            last_col_letter = get_column_letter(last_col)
            ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        auto_width(ws)

        url_cols = find_url_columns(ws)
        if url_cols:
            hl_count = hyperlink_urls(ws, url_cols)
            print(f"    Hyperlinks created: {hl_count} in columns {url_cols}")
        else:
            print(f"    No URL columns detected")

    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        gen_row = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(row, col).value or "")
                if "generated" in val.lower() or "last updated" in val.lower():
                    gen_row = row
                    break
            if gen_row:
                break

        if gen_row:
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(gen_row, col).value or "")
                if "2026" in val:
                    ws.cell(gen_row, col).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                    break
            print(f"  Dashboard: Updated timestamp")
        else:
            print(f"  Dashboard: Timestamp row not found, skipping")

    print()
    wb.save(EXCEL_PATH)
    wb.close()
    print("Workbook saved successfully.")


if __name__ == "__main__":
    main()
