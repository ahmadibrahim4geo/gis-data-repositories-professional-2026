"""
Update Free_GIS_Data sheet from source website.

This script scrapes https://freegisdata.rtwilson.com/ and updates the
Excel workbook's Free_GIS_Data sheet. It creates a backup before saving.

Usage:
    python scripts/update_free_gis_data.py

Note: If the source website changes its structure, this script may need
updating. Network failures are handled gracefully.
"""

import sys
from datetime import date, datetime
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
EXCEL_PATH = DATA_DIR / "GIS_Data_Repositories_Professional_2026.xlsx"
SHEET_NAME = "Free_GIS_Data"
SOURCE_URL = "https://freegisdata.rtwilson.com/"

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; GIS-Data-Catalog/1.0)"}


def update():
    print(f"Free GIS Data Updater")
    print(f"Source: {SOURCE_URL}")
    print(f"Date: {datetime.now().isoformat()}")
    print()

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    try:
        import requests
    except ImportError:
        print("ERROR: requests is required. Install with: pip install requests")
        sys.exit(1)

    print("Fetching source page...")
    try:
        resp = requests.get(SOURCE_URL, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        html = resp.text
        print(f"Got {len(html):,} bytes from source")
    except requests.RequestException as e:
        print(f"Network error: {e}")
        print("Cannot update from web. The existing data is preserved.")
        print("Check your internet connection or try again later.")
        return

    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    links = []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True)
        if href.startswith("http") and text:
            links.append((text, href))

    print(f"Found {len(links)} external links on the source page")

    backup_path = DATA_DIR / f"GIS_Data_Repositories_Professional_2026_backup_{date.today().isoformat()}.xlsx"
    import shutil
    shutil.copy2(EXCEL_PATH, backup_path)
    print(f"Backup saved: {backup_path.name}")

    wb = openpyxl.load_workbook(EXCEL_PATH)

    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet '{SHEET_NAME}' not found in workbook")
        wb.close()
        return

    ws = wb[SHEET_NAME]

    ws.cell(row=2, column=1).value = f"Source: {SOURCE_URL}"
    ws.cell(row=2, column=4).value = f"Last checked: {date.today().isoformat()}"
    ws.cell(row=2, column=7).value = f"Links found: {len(links)}"

    wb.save(EXCEL_PATH)
    wb.close()

    print(f"Updated. Links found on source page: {len(links)}")
    print(f"Manual review of the Free_GIS_Data sheet is recommended to")
    print(f"merge new links and verify existing entries.")


if __name__ == "__main__":
    update()
